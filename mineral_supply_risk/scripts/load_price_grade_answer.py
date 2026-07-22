# -*- coding: utf-8 -*-
"""KOMIS 가격기준 주간 이격률 모니터링 xlsx → warehouse 적재: fact_diagnosis_answer.

배경(2026-07-16, 사용자 지시): documents/2차_데이타/3. 학습 및 검증용/1. 학습용 참고자료/
「1. 주간가격이격률모니터링_코미스가격기준 (1).xlsx」의 '등급모니터링'(광종별 주간 등급)+
'가격DB'(동일 그리드 주간 가격)+'이격률'(동일 그리드 연속형 이격률 수치) 시트를 적재.

**용도 정정(2026-07-16, 같은 날 재확인)**: 최초에는 이 KOMIS 등급을 진단모델의 정답셋(target)
으로 쓰려 했으나, 사용자 확인 결과 **정답(target)은 기존 프로젝트의 4단계 수급위기 경보
(out_diagnosis_alert / crisis_index, teacher_supply_demand 기반)로 되돌리고, 이 KOMIS
가격이격률 데이터(연속형 deviation_rate, 등급 grade는 참고용)는 기존 진단모델의 신규 피처로
사용한다.** grade/grade_ord는 참고용으로 계속 보존(향후 다른 분석에 재사용 가능하도록
삭제하지 않음 — artifact-provenance-policy 원칙).

등급 정의('참고사항' 시트, 2026-07-15 일루넥스 확인, 참고용): 이격률(가격의 과거 평균 대비
표준편차 배수)의 상방(+) 이탈만 감지하는 3단계 규칙 — 정상: 이격률 < 평균+1σ / 관심:
평균+1σ ≤ 이격률 < 평균+2σ / 주의경계심각: 이격률 ≥ 평균+2σ.

컬럼 선택: fact_price(load_komis_xlsx.py PRICE_COLS)와 동일한 5광종·동일 가격기준으로 한정
(정합성 유지) — CU=동/LME CASH, NI=니켈/LME CASH, CO=코발트/LME CASH, LI=탄산리튬/99.5%min
CIF China, REE=산화네오디뮴/99.5%min FOB China.

실행: MSR_DB=<warehouse> python -m scripts.load_price_grade_answer
"""
from __future__ import annotations
import argparse

import duckdb
import openpyxl
import pandas as pd

from msr.config import DB_PATH

XLSX = ("/home/nuri/dev/git/ws/mine_ws/komir/documents/2차_데이타/3. 학습 및 검증용/"
        "1. 학습용 참고자료/1. 주간가격이격률모니터링_코미스가격기준 (1).xlsx")

# (광종명, 가격기준 접두) → (commodity_code, series_label) — fact_price.PRICE_COLS와 동일 정의
COLS = [
    ("동", "LME CASH", "CU"),
    ("니켈", "LME CASH", "NI"),
    ("코발트", "LME CASH", "CO"),
    ("탄산리튬", "99.5%min", "LI"),
    ("산화네오디뮴", "99.5%min", "REE"),
]
GRADE_ORD = {"정상": 0, "관심": 1, "주의경계심각": 2}


def _find_col(name_row, basis_row, kname, kbasis):
    for i in range(3, len(name_row) + 1):
        n = name_row[i - 1]
        b = basis_row[i - 1]
        if n and str(n).strip() == kname and b and str(b).startswith(kbasis):
            return i
    return None


def load() -> pd.DataFrame:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws_grade, ws_price = wb["등급모니터링"], wb["가격DB"]
    ws_dev = wb["이격률"]  # 등급모니터링과 행·열 정렬 동일(동일 워크북 레이아웃, 실측 확인됨)

    name_row = [ws_grade.cell(row=1, column=c).value for c in range(1, ws_grade.max_column + 1)]
    basis_row = [ws_grade.cell(row=2, column=c).value for c in range(1, ws_grade.max_column + 1)]

    # 실데이터 행 범위: 기준일이 'YYYYMMDD' 문자열인 행만(꼬리 빈 수식행은 time 타입 → 제외)
    last_row = 2
    for r in range(3, ws_grade.max_row + 1):
        v = ws_grade.cell(row=r, column=2).value
        if isinstance(v, str) and v.isdigit() and len(v) == 8:
            last_row = r
        else:
            break

    rows = []
    for kname, kbasis, cc in COLS:
        col = _find_col(name_row, basis_row, kname, kbasis)
        if col is None:
            print(f"  [warn] 컬럼 미발견: {kname}/{kbasis}")
            continue
        series_label = ws_grade.cell(row=2, column=col).value
        for r in range(3, last_row + 1):
            d = ws_grade.cell(row=r, column=2).value
            grade = ws_grade.cell(row=r, column=col).value
            price = ws_price.cell(row=r, column=col).value  # 가격DB는 등급모니터링과 동일 행·열 정렬
            dev = ws_dev.cell(row=r, column=col).value       # 이격률(연속형)도 동일 행·열 정렬
            if grade not in GRADE_ORD:
                continue  # #DIV/0!(표준편차 계산 불가 초기구간) 또는 None(등급 미산정) 제외
            obs_date = pd.to_datetime(d, format="%Y%m%d", errors="coerce")
            if pd.isna(obs_date):
                continue
            rows.append(dict(
                commodity_code=cc, indicator="PRICE_DEVIATION_GRADE", freq="W",
                obs_date=obs_date.date(), grade=grade, grade_ord=GRADE_ORD[grade],
                price=float(price) if isinstance(price, (int, float)) else None,
                deviation_rate=float(dev) if isinstance(dev, (int, float)) else None,
                series_label=series_label, src="KOMIS_GRADE_MONITOR",
            ))
    out = pd.DataFrame(rows)
    out = out.drop_duplicates(["commodity_code", "indicator", "obs_date"], keep="last")
    return out


def run(db: str | None = None) -> dict:
    df = load()
    con = duckdb.connect(db or DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS fact_diagnosis_answer (
            commodity_code VARCHAR NOT NULL,
            indicator VARCHAR NOT NULL,
            freq VARCHAR,
            obs_date DATE NOT NULL,
            grade VARCHAR,
            grade_ord INTEGER,
            price DECIMAL(20,4),
            deviation_rate DECIMAL(20,6),
            series_label VARCHAR,
            src VARCHAR,
            PRIMARY KEY (commodity_code, indicator, obs_date)
        )
    """)
    existing_cols = {r[1] for r in con.execute("PRAGMA table_info('fact_diagnosis_answer')").fetchall()}
    if "deviation_rate" not in existing_cols:
        con.execute("ALTER TABLE fact_diagnosis_answer ADD COLUMN deviation_rate DECIMAL(20,6)")
    con.register("_g", df)
    con.execute("DELETE FROM fact_diagnosis_answer WHERE src='KOMIS_GRADE_MONITOR'")
    # 명시적 컬럼 리스트 사용(테이블 실제 컬럼 순서가 ALTER TABLE 이력에 따라 CREATE TABLE
    # 선언 순서와 달라질 수 있어 위치기반 INSERT는 위험 — 2026-07-16 실측으로 확인된 버그).
    con.execute("""INSERT INTO fact_diagnosis_answer
        (commodity_code, indicator, freq, obs_date, grade, grade_ord, price,
         deviation_rate, series_label, src)
        SELECT commodity_code, indicator, freq, obs_date, grade, grade_ord, price,
               deviation_rate, series_label, src FROM _g""")
    con.execute("CHECKPOINT")
    chk = con.execute("""
        SELECT commodity_code, min(obs_date), max(obs_date), count(*),
               sum(CASE WHEN grade='정상' THEN 1 ELSE 0 END) AS n_정상,
               sum(CASE WHEN grade='관심' THEN 1 ELSE 0 END) AS n_관심,
               sum(CASE WHEN grade='주의경계심각' THEN 1 ELSE 0 END) AS n_주의경계심각
        FROM fact_diagnosis_answer WHERE src='KOMIS_GRADE_MONITOR'
        GROUP BY 1 ORDER BY 1""").fetchall()
    con.close()
    print(f"[price_grade_answer] fact_diagnosis_answer {len(df)}행 적재")
    for r in chk:
        print("  ", r)
    return {"rows": len(df)}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=None)
    run(ap.parse_args().db)
