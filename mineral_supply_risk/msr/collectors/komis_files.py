# -*- coding: utf-8 -*-
"""
핵심광물 프로젝트 - 보유(로컬 xlsx/csv) 데이터 -> DuckDB 적재 **독립 스크립트**.

⚠️ 스키마 계층 주의:
  - 이 로더는 **canonical 스키마 `data/raw/00_schema.sql`**(fact_trade·dim_commodity·
    dim_series 등 정규화 테이블)에 적재한다. 모델 마트(features/marts.py)가 기대하는 스키마.
  - 도커 수집 파이프라인(scripts.run)이 적용하는 `db/schema_core.sql`(raw_customs_*·
    fact_trade_annual 계열)과는 **별개 스키마**다. 아직 두 계층이 통합돼 있지 않다.
  - 파이프라인(scripts.run/compose)에 연결돼 있지 않은 standalone 도구. 로컬 원본 파일을
    한 번에 canonical DuckDB로 적재할 때 수동 실행한다.

사용법:
    python -m msr.collectors.komis_files --data-root "<데이터폴더>" --db "minerals.duckdb"
의존성: duckdb, pandas, openpyxl
"""
import argparse, os, glob, re, datetime as dt
import pandas as pd
import duckdb

# canonical 스키마 파일(프로젝트 data/raw/00_schema.sql) 절대경로
_SCHEMA_DEFAULT = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "00_schema.sql"))

def _bulk(con, table, cols, rows, del_where=None):
    """멱등 일괄삽입: del_where로 기존행 삭제 후 DataFrame로 고속 INSERT"""
    if del_where:
        con.execute(f"DELETE FROM {table} WHERE {del_where}")
    if not rows:
        return 0
    df = pd.DataFrame(rows, columns=cols)
    con.register("_bulk_df", df)
    con.execute(f"INSERT INTO {table} SELECT * FROM _bulk_df")
    con.unregister("_bulk_df")
    return len(rows)

PRICE_COLS=["commodity_code","raw_name","price_basis","unit","freq","obs_date","value","source"]
INV_COLS=["commodity_code","raw_name","obs_date","lme_stock","stock_chg","stock_chg_pct","unit","source"]
TRADE_COLS=["year","commodity_code","raw_name","hscode","item_name","country","import_value_usd","import_weight_kg","export_value_usd","export_weight_kg","flow_lv1","flow_lv2","classification","source"]
PR_COLS=["source_tag","commodity_code","raw_name","country","metric_type","year","value","unit","source"]
IND_COLS=["commodity_code","raw_name","obs_date","indicator_type","value","source"]
SER_COLS=["series_code","obs_date","value","unit","source"]

# ---- 5대 핵심광물 사전 ----
CORE = [
    ("CU","동","Copper",True,"비철금속"),
    ("NI","니켈","Nickel",True,"비철금속"),
    ("LI","리튬","Lithium",True,"배터리"),
    ("CO","코발트","Cobalt",True,"배터리"),
    ("REE","희토류","Rare earths",True,"희토류"),
]
# 원천 표기 -> 표준코드 (필요 시 확장; HS코드 매핑은 별도 검증 과업)
NAME2CODE = {
    "동":"CU","Copper":"CU","구리":"CU",
    "니켈":"NI","Nickel":"NI",
    "리튬":"LI","탄산리튬":"LI","수산화리튬":"LI","Lithium":"LI",
    "코발트":"CO","Cobalt":"CO",
    "희토류":"REE","Rare earths":"REE","네오디뮴":"REE","란탄":"REE",
    "디스프로슘":"REE","터븀":"REE","세륨":"REE",
}
def code_of(name):
    if name is None: return None
    return NAME2CODE.get(str(name).strip())

def to_date(v):
    """YYYYMMDD / YYYY/MM/DD / datetime -> date"""
    if v is None or (isinstance(v,float) and pd.isna(v)): return None
    if isinstance(v,(dt.datetime,dt.date)): return pd.Timestamp(v).date()
    s=str(v).strip()
    qm=re.match(r"(\d{4})[/\-]([1-4])Q",s,re.I)        # 2016/1Q
    if qm: return dt.date(int(qm.group(1)),(int(qm.group(2))-1)*3+1,1)
    ym=re.match(r"^(\d{4})[/\-](\d{1,2})$",s)          # 2016/01
    if ym: return dt.date(int(ym.group(1)),int(ym.group(2)),1)
    for fmt in ("%Y%m%d","%Y/%m/%d","%Y-%m-%d"):
        try: return dt.datetime.strptime(s[:10] if "-" in s or "/" in s else s,fmt).date()
        except: pass
    return None

def num(v):
    try:
        f=float(str(v).replace(",",""))
        return f
    except: return None

# =====================================================================
def load_hs_map(con, root):
    """HS코드분류__최종.xlsx 종합분류 -> dim_hs_commodity (검증완료 권위 매핑)"""
    import openpyxl
    f=glob.glob(os.path.join(root,"**","HS*분류*.xlsx"),recursive=True)
    if not f: print("  [skip] HS분류 없음"); return {}
    wb=openpyxl.load_workbook(f[0],data_only=True,read_only=True)
    ws=wb["종합분류"]; rows=list(ws.iter_rows(values_only=True)); wb.close()
    NORM={"구리":"동"}; CODE={"동":"CU","니켈":"NI","리튬":"LI","코발트":"CO","희토류":"REE"}
    recs=[]; hs2code={}
    seen=set()
    for r in rows[1:]:
        lv1=r[10]; hs10=r[13]
        if hs10 is None or lv1 is None: continue
        hs=str(hs10).strip(); ko=NORM.get(str(lv1).strip(),str(lv1).strip())
        if hs in seen: continue
        seen.add(hs)
        cc=CODE.get(ko)
        recs.append((hs,ko,cc,ko in CODE))
        if cc: hs2code[hs]=cc
    con.execute("DELETE FROM dim_hs_commodity")
    con.executemany("INSERT INTO dim_hs_commodity VALUES (?,?,?,?)",recs)
    print(f"  dim_hs_commodity={len(recs)} (core5 HS={sum(1 for x in recs if x[3])})")
    return hs2code

def load_dims(con):
    con.execute("DELETE FROM dim_commodity")
    con.executemany("INSERT INTO dim_commodity VALUES (?,?,?,?,?)", CORE)
    rows=[("MANUAL",k,v) for k,v in NAME2CODE.items()]
    con.execute("DELETE FROM dim_commodity_map")
    con.executemany("INSERT INTO dim_commodity_map VALUES (?,?,?)", rows)
    print(f"  dim_commodity={len(CORE)}, dim_commodity_map={len(rows)}")

# ---- 1) 공급망통계 '주간 평균' : LME CASH / 3개월 (멀티헤더 0~4) ----
def load_supply_weekly(con, root):
    import openpyxl
    f=glob.glob(os.path.join(root,"**","0. KOMIS*공급망 통계*xlsx"),recursive=True)
    if not f: print("  [skip] 공급망통계 없음"); return
    wb=openpyxl.load_workbook(f[0],data_only=True,read_only=True)
    ws=wb["주간 평균"]
    rows=list(ws.iter_rows(values_only=True)); wb.close()
    commodity=rows[1]; basis=rows[3]; unit=rows[4]
    recs=[]
    for r in rows[5:]:
        d=to_date(r[0])
        if not d: continue
        for c in range(1,len(commodity)):
            val=num(r[c])
            if val is None: continue
            b=str(basis[c]); pb="LME_CASH" if "CASH" in b.upper() else ("LME_3M" if "3" in b else b)
            recs.append((code_of(commodity[c]),commodity[c],pb,unit[c],"W",d,val,"KOMIS_SUPPLY"))
    n=_bulk(con,"fact_price",PRICE_COLS,recs,del_where="source='KOMIS_SUPPLY'")
    print(f"  fact_price(공급망 주간 Cash/3M) += {n}")

# ---- 1b/2) 주간가격및재고량_*.xlsx : 기준가격 + LME재고 ----
def load_weekly_price_inv(con, root):
    import openpyxl
    files=glob.glob(os.path.join(root,"**","1. 주간가격및재고량_*.xlsx"),recursive=True)
    pr=[]; iv=[]
    for f in files:
        m=re.search(r"주간가격및재고량_(.+)\.xlsx",os.path.basename(f))
        name=m.group(1) if m else None; cc=code_of(name)
        wb=openpyxl.load_workbook(f,data_only=True)  # read_only시 컬럼 잘림 -> 일반 모드
        ws=wb.active
        rows=[[ws.cell(row=rr,column=cc).value for cc in range(1,10)] for rr in range(1,ws.max_row+1)]
        wb.close()
        # 헤더는 3행(index2): 기준일,기준가격,최저가,최고가,...,LME재고량,LME전일대비등락가,LME전일대비등락비율
        for r in rows[3:]:
            d=to_date(r[0])
            if not d: continue
            if num(r[1]) is not None:
                pr.append((cc,name,"REF","USD/mt","W",d,num(r[1]),"KOMIS_WEEKLY"))
            if len(r)>6 and num(r[6]) is not None:
                iv.append((cc,name,d,num(r[6]),num(r[7]) if len(r)>7 else None,
                           num(r[8]) if len(r)>8 else None,"mt","KOMIS_WEEKLY"))
    np_=_bulk(con,"fact_price",PRICE_COLS,pr,del_where="source='KOMIS_WEEKLY'")
    ni_=_bulk(con,"fact_inventory",INV_COLS,iv,del_where="source='KOMIS_WEEKLY'")
    print(f"  fact_price(주간 기준가) += {np_} | fact_inventory += {ni_}")

# ---- 3) 관세청 수출입DB (연간·국가별) ----

def load_supply_price_ym(con, root):
    """공급망통계 '핵심광물가격' 시트(41종, 연 2021~25 + 월 2026) -> fact_price(freq Y/M)"""
    import openpyxl, datetime as dt
    f=glob.glob(os.path.join(root,"**","0. KOMIS*공급망 통계*xlsx"),recursive=True)
    if not f: print("  [skip] 공급망통계 없음"); return
    CODE={"니켈":"NI","동":"CU","코발트":"CO","탄산리튬":"LI",
          "산화네오디뮴":"REE","산화디스프로슘":"REE","산화란탄":"REE","산화세륨":"REE",
          "산화터븀":"REE","산화이트륨":"REE","산화스칸듐":"REE"}
    wb=openpyxl.load_workbook(f[0],data_only=True,read_only=True)
    ws=wb["핵심광물가격(2026.5월 기준)"]; rows=list(ws.iter_rows(values_only=True)); wb.close()
    hdr=rows[1]; colmap={}
    for c in range(4,22):
        lab=str(hdr[c]) if c<len(hdr) and hdr[c] else ""
        m=re.match(r"(\d{4})년\s*(\d{1,2})월",lab)
        if m: colmap[c]=("M",dt.date(int(m.group(1)),int(m.group(2)),1)); continue
        m=re.match(r"(\d{4})년(?!.*월)",lab)
        if m and "누적" not in lab: colmap[c]=("Y",dt.date(int(m.group(1)),1,1))
    recs=[]
    for r in rows[2:]:
        if not r or len(r)<4 or not r[1]: continue
        name=str(r[1]).strip(); basis=str(r[2]) if r[2] else None; unit=str(r[3]) if r[3] else None
        for c,(freq,d) in colmap.items():
            v=num(r[c]) if c<len(r) else None
            if v is None or v<=0: continue
            recs.append((CODE.get(name),name,basis,unit,freq,d,v,"KOMIS_SUPPLY_YM"))
    n=_bulk(con,"fact_price",PRICE_COLS,recs,del_where="source='KOMIS_SUPPLY_YM'")
    print(f"  fact_price(핵심광물가격 연·월) += {n}")

def load_customs(con, root, hs2code=None, limit=None):
    import openpyxl
    f=glob.glob(os.path.join(root,"**","(국내)관세청*수출입DB*xlsx"),recursive=True)
    if not f: print("  [skip] 관세청 없음"); return
    wb=openpyxl.load_workbook(f[0],data_only=True,read_only=True)
    ws=wb["2013~2025_수출입DB_260521"]
    it=ws.iter_rows(values_only=True)
    hdr=[str(c).strip() if c is not None else "" for c in next(it)]
    idx={name:i for i,name in enumerate(hdr)}
    def g(r,name): 
        i=idx.get(name); return r[i] if i is not None and i<len(r) else None
    recs=[]
    for n,r in enumerate(it):
        if limit and n>=limit: break
        hs=str(g(r,"HSCODE") or "").strip()
        cc=(hs2code or {}).get(hs) or code_of(g(r,"광종명"))
        recs.append((int(num(g(r,"연도")) or 0) or None, cc, g(r,"광종명"),
                     g(r,"HSCODE"), g(r,"품목명"), g(r,"국가명"),
                     num(g(r,"수입액($)")), num(g(r,"수입중량(kg)")),
                     num(g(r,"수출액($)")), num(g(r,"수출중량(kg)")),
                     g(r,"물질흐름_Lv1"), g(r,"물질흐름_Lv2"), g(r,"분류"), "CUSTOMS"))
    wb.close()
    n=_bulk(con,"fact_trade",TRADE_COLS,recs,del_where="source='CUSTOMS'")
    print(f"  fact_trade += {n}")

# ---- 4) USGS 생산·매장량 (피벗전용데이터 -> long) ----
def load_usgs(con, root):
    f=glob.glob(os.path.join(root,"**","0. USGS_엑셀정리본*xlsx"),recursive=True)
    if not f: print("  [skip] USGS 없음"); return
    df=pd.read_excel(f[0],sheet_name="피벗전용데이터")
    df.columns=[str(c).replace("\n"," ").strip() for c in df.columns]
    # 생산 컬럼 후보
    EXC=("점유","국가","순","계","비중","share","rank","%")
    col_prod=[c for c in df.columns if "생산" in c and not any(e in c for e in EXC)]
    col_res =[c for c in df.columns if "가채매장량" in c and not any(e in c for e in EXC)]
    def yr(c):
        m=re.search(r"20\d{2}",c); return int(m.group()) if m else None
    recs=[]
    for _,x in df.iterrows():
        src=x.get("SOURCE"); com=x.get("COMMODITY"); ctry=x.get("COUNTRY"); unit=x.get("UNIT")
        cc=code_of(com)
        for c in col_prod:
            v=num(x.get(c))
            if v is not None: recs.append((src,cc,com,ctry,"PRODUCTION",yr(c),v,unit,"USGS"))
        for c in col_res:
            v=num(x.get(c))
            if v is not None: recs.append((src,cc,com,ctry,"RESERVE",yr(c),v,unit,"USGS"))
    n=_bulk(con,"fact_production_reserve",PR_COLS,recs,del_where="source='USGS'")
    print(f"  fact_production_reserve += {n}")

# ---- 5) KOMIS 지표 (수급동향/시장동향/광물종합, 월간 wide -> long) ----
def load_komis_indicators(con, root):
    targets=[("수급동향지표","SUPPLY_DEMAND"),("시장동향지표","MARKET"),("광물종합지수","COMPOSITE")]
    con.execute("DELETE FROM fact_indicator WHERE source LIKE 'KOMIS_%'")
    total=0
    for fname,itype in targets:
        f=glob.glob(os.path.join(root,"**",fname+".xlsx"),recursive=True)
        if not f: continue
        import openpyxl
        wb=openpyxl.load_workbook(f[0],data_only=True,read_only=True)
        # 지표 시트 + '가격' 시트. '가격'은 수급동향지표에서만 1회 적재(이중적재 방지)
        for sheet in wb.sheetnames:
            if sheet=="가격" and fname!="수급동향지표":
                continue
            ws=wb[sheet]; rows=list(ws.iter_rows(values_only=True))
            # 헤더행: '구분'이 들어있는 행
            hidx=next((i for i,r in enumerate(rows) if r and any(str(c).strip()=="구분" for c in r if c)),None)
            if hidx is None: continue
            hdr=rows[hidx]; commo_cols={c:hdr[c] for c in range(len(hdr)) if hdr[c] and str(hdr[c]).strip() not in("","구분")}
            it = "PRICE" if sheet=="가격" else itype
            recs=[]
            for r in rows[hidx+1:]:
                d=to_date(r[1] if len(r)>1 else None)
                if not d: continue
                for c,name in commo_cols.items():
                    v=num(r[c]) if c<len(r) else None
                    if v is None: continue
                    recs.append((code_of(name),name,d,it,v,"KOMIS_"+fname))
            if recs:
                total+=_bulk(con,"fact_indicator",IND_COLS,recs); total=total
        wb.close()
    print(f"  fact_indicator += {total}")

# ---- 6) 거시·지수·환율 CSV (CP949, long) ----
def load_series_csv(con, root):
    files=sorted(set(os.path.realpath(p) for p in glob.glob(os.path.join(root,"**","*.csv"),recursive=True)))
    recs=[]; meta_map={}
    for f in files:
        base=os.path.basename(f)
        code=re.sub(r"^\d+\.\s*","",base).replace(".csv","").strip()
        try:
            lines=open(f,encoding="cp949").read().splitlines()
        except:
            lines=open(f,encoding="utf-8",errors="ignore").read().splitlines()
        title=lines[0] if lines else code
        # 헤더행: '날짜'로 시작하는 행
        hidx=next((i for i,l in enumerate(lines) if l.startswith("날짜")),None)
        if hidx is None: continue
        unit=""
        for l in lines[hidx+1:]:
            p=l.split(",")
            if len(p)<2: continue
            d=to_date(p[0]); v=num(p[1])
            if d is None or v is None: continue
            if len(p)>=3: unit=p[2]
            recs.append((code,d,v,unit,"CSV"))
        fr="M" if "월간" in title else ("Q" if "분기" in title else ("D" if "일간" in title else "W"))
        meta_map[code]=(code,title,unit,fr,"CSV")
    if True: _bulk(con,"fact_series",SER_COLS,recs,del_where="source='CSV'")
    if meta_map:
        con.execute("DELETE FROM dim_series")
        con.executemany("INSERT INTO dim_series VALUES (?,?,?,?,?)",list(meta_map.values()))
    print(f"  fact_series += {len(recs)} | dim_series += {len(meta_map)}")

# =====================================================================
def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--data-root",required=True)
    ap.add_argument("--db",default="minerals.duckdb")
    ap.add_argument("--schema",default=_SCHEMA_DEFAULT)
    a=ap.parse_args()
    con=duckdb.connect(a.db)
    con.execute(open(a.schema,encoding="utf-8").read())
    print("[1] dims");          load_dims(con)
    print("[1b] HS->광종 매핑");  hs2code=load_hs_map(con,a.data_root)
    print("[2] 공급망 주간 Cash/3M"); load_supply_weekly(con,a.data_root)
    print("[3] 주간 기준가/재고");    load_weekly_price_inv(con,a.data_root)
    print("[3b] 핵심광물가격 연·월"); load_supply_price_ym(con,a.data_root)
    print("[4] 관세청 교역");        load_customs(con,a.data_root,hs2code=hs2code)
    print("[5] USGS 생산/매장");     load_usgs(con,a.data_root)
    print("[6] KOMIS 지표");        load_komis_indicators(con,a.data_root)
    print("[7] 거시/지수/환율 CSV"); load_series_csv(con,a.data_root)
    print("\n=== 적재 요약 ===")
    for t in ["fact_price","fact_inventory","fact_trade","fact_production_reserve","fact_indicator","fact_series"]:
        n=con.execute(f"SELECT count(*) FROM {t}").fetchone()[0]
        print(f"  {t:28s}: {n:,}")
    con.close()

if __name__=="__main__":
    main()
